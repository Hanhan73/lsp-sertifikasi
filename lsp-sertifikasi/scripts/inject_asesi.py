#!/usr/bin/env python3
"""
inject_asesi.py - Inject nama asesi ke template observasi/portofolio Excel
Otomatis mencari kolom yang headernya "Nama" dan mengisi nama di bawahnya.

Usage:
    python3 inject_asesi.py <input_file> <output_file> <nama1> [<nama2> ...]

Example:
    python3 inject_asesi.py template.xlsm output.xlsm "Budi Santoso" "Ani Rahayu" "Citra Dewi"
"""

import sys
import shutil
import openpyxl

SKIP_SHEETS = {'Cover', 'Asesor', 'Berita Acara', 'Hasil Asesmen'}


def find_nama_header(ws, max_row=5, max_col=15):
    """Cari cell yang isinya 'Nama' (case-insensitive) dalam beberapa baris pertama."""
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == 'nama':
                return cell.row, cell.column
    return None, None


def detect_target_sheets(wb):
    """
    Deteksi sheet mana saja yang perlu diisi nama.
    - Kalau sheet ke-2 dst punya formula referensi ke sheet pertama -> cukup isi sheet pertama.
    - Kalau tidak -> isi semua sheet data.
    """
    data_sheets = [s for s in wb.sheetnames if s not in SKIP_SHEETS]
    if not data_sheets:
        return []

    first_sheet = data_sheets[0]
    if len(data_sheets) > 1:
        ws2 = wb[data_sheets[1]]
        hrow, hcol = find_nama_header(ws2)
        if hrow and hcol:
            cell_val = ws2.cell(row=hrow + 1, column=hcol).value
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('=') and first_sheet in cell_val:
                return [first_sheet]

    return data_sheets


def inject_names(input_path, output_path, names):
    shutil.copy2(input_path, output_path)

    wb = openpyxl.load_workbook(output_path, keep_vba=True)
    target_sheets = detect_target_sheets(wb)

    if not target_sheets:
        print("ERROR: Tidak ada sheet data yang ditemukan.")
        sys.exit(1)

    injected_sheets = []
    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        header_row, nama_col = find_nama_header(ws)
        if header_row is None:
            print(f"  SKIP [{sheet_name}]: header 'Nama' tidak ditemukan.")
            continue

        data_start_row = header_row + 2
        rows_needed = set(range(data_start_row, data_start_row + len(names)))

        # Unmerge merged cells di kolom nama agar bisa ditulis
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_col <= nama_col <= merge_range.max_col:
                if rows_needed & set(range(merge_range.min_row, merge_range.max_row + 1)):
                    ws.unmerge_cells(str(merge_range))

        for i, nama in enumerate(names):
            ws.cell(row=data_start_row + i, column=nama_col).value = nama

        injected_sheets.append(sheet_name)

    wb.save(output_path)
    print(f"OK: {len(names)} asesi diinjeksi ke sheet: {', '.join(injected_sheets)}")


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(f"Usage: {sys.argv[0]} <input_file> <output_file> <nama1> [<nama2> ...]")
        sys.exit(1)

    inject_names(sys.argv[1], sys.argv[2], sys.argv[3:])