#!/usr/bin/env python3
"""
scripts/parse_berita_acara.py
Baca sheet 'Berita Acara' dari file Excel penilaian (.xlsm/.xlsx).
Extract nama peserta + rekomendasi K/BK, output sebagai JSON.

Usage: python3 parse_berita_acara.py '<filepath>'
"""

import sys
import json

try:
    import openpyxl
except ImportError:
    print(json.dumps({'error': 'openpyxl not installed'}))
    sys.exit(1)


def parse(filepath: str) -> dict:
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
    except Exception as e:
        return {'error': f'Gagal membuka file: {e}'}

    if 'Berita Acara' not in wb.sheetnames:
        return {'error': "Sheet 'Berita Acara' tidak ditemukan di file ini"}

    ws = wb['Berita Acara']

    # Cari teks pembuka "Pada tanggal..."
    tanggal_raw = None
    for r in range(1, 20):
        cell = ws.cell(r, 3).value
        if cell and 'Pada tanggal' in str(cell):
            tanggal_raw = str(cell).strip()
            break

    # Deteksi mode rekomendasi:
    # - 2 kolom: row18 col K = 'BK'  → K di col J, BK di col K (centang/isi)
    # - 1 kolom: col J langsung isi 'K' atau 'BK'
    row18_k = ws.cell(18, 11).value
    two_col_mode = str(row18_k or '').strip().upper() == 'BK'

    peserta = []
    for r in range(19, ws.max_row + 1):
        no_val   = ws.cell(r, 3).value
        nama_val = ws.cell(r, 4).value

        # Stop jika nomor bukan digit atau nama kosong/placeholder
        if not no_val or not str(no_val).strip().isdigit():
            break
        if not nama_val or str(nama_val).strip() in ['0', '', '-']:
            break

        nama = str(nama_val).strip()
        rek_j = ws.cell(r, 10).value  # col J
        rek_k = ws.cell(r, 11).value  # col K

        if two_col_mode:
            # Ada isian di col J → K, ada isian di col K → BK
            val_j = str(rek_j or '').strip().upper()
            val_k = str(rek_k or '').strip().upper()
            check_vals = {'K', 'V', '✓', 'X', '1', 'YA', 'Y'}
            if val_j in check_vals:
                rek = 'K'
            elif val_k in check_vals:
                rek = 'BK'
            else:
                rek = None
        else:
            # Single col J: langsung 'K' atau 'BK'
            val = str(rek_j or '').strip().upper()
            rek = val if val in ['K', 'BK'] else None

        peserta.append({'nama': nama, 'rekomendasi': rek})

    return {
        'tanggal_raw': tanggal_raw,
        'peserta': peserta,
        'total': len(peserta),
        'filled': sum(1 for p in peserta if p['rekomendasi'] is not None),
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'Usage: parse_berita_acara.py <filepath>'}))
        sys.exit(1)

    result = parse(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False))